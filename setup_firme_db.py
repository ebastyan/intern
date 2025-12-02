# -*- coding: utf-8 -*-
"""
PAJU - Firme (Vanzari) Database Setup and Import
=================================================
Ez a script:
1. Letrehozza a firme tablakat NeonDB-ben
2. Importalja az osszes adatot a 3 XLSX fajlbol (2022, 2023, 2024)
3. Cegenkenti es hulladek tipus szerinti osszesiteseket ment

Tablak:
- firme: Cegek (vevok/ugyfelek)
- vanzari: Eladasok (aviz szinten)
- sumar_firme: Havi osszesites cegenkent
- sumar_deseuri: Havi osszesites hulladek tipusonkent
- transporturi_firme: Szallitasok
"""

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
import os
import re
import sys
from datetime import datetime, date
from decimal import Decimal

# UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

# NeonDB connection
DATABASE_URL = "postgresql://neondb_owner:npg_L2AyrcXul8km@ep-ancient-firefly-a47vk6i8-pooler.us-east-1.aws.neon.tech/neondb?sslmode=require"
BASE_PATH = r'C:\Users\INTEL\Desktop\paju\situatie vanzari'

# Month name to number mapping
MONTH_MAP = {
    'IANUARIE': 1, 'FEBRUARIE': 2, 'MARTIE': 3, 'APRILIE': 4,
    'MAI': 5, 'IUNIE': 6, 'IULIE': 7, 'AUGUST': 8,
    'SEPTEMBRIE': 9, 'OCTOMBRIE': 10, 'NOIEMBRIE': 11, 'DECEMBRIE': 12
}

def get_connection():
    return psycopg2.connect(DATABASE_URL)

def normalize_string(s):
    """Normalize string: strip whitespace"""
    if s is None:
        return None
    s = str(s).strip()
    s = re.sub(r'\s+', ' ', s)
    return s if s else None

def normalize_company_name(name):
    """Normalize company name for matching"""
    if not name:
        return None
    name = normalize_string(name)
    if not name:
        return None
    # Remove common suffixes
    name = re.sub(r'\s+(S\.?R\.?L\.?|SRL|SA|S\.A\.|BV|DOO|SPOLKA.*|GMBH|LTD)\.?\s*$', '', name, flags=re.IGNORECASE)
    return name.strip().upper()

def parse_date(date_val):
    """Parse date from various formats"""
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val.date()
    if isinstance(date_val, date):
        return date_val

    date_str = str(date_val).strip()
    formats = ['%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y']
    for fmt in formats:
        try:
            return datetime.strptime(date_str.split()[0], fmt).date()
        except:
            pass
    return None

def safe_float(val):
    """Safely convert to float"""
    if val is None:
        return None
    try:
        return float(val)
    except:
        return None

def safe_decimal(val):
    """Safely convert to Decimal for DB"""
    if val is None:
        return None
    try:
        return round(float(val), 4)
    except:
        return None

# =============================================================================
# MAIN SCRIPT
# =============================================================================

print("=" * 70)
print("PAJU - Firme Database Setup and Import")
print("=" * 70)

conn = get_connection()
cur = conn.cursor()

# =============================================================================
# CREATE TABLES
# =============================================================================
print("\n[1] Creare tabele pentru Firme...")

cur.execute("""
-- Drop existing firme tables (don't touch persoane fizice tables!)
DROP TABLE IF EXISTS transporturi_firme CASCADE;
DROP TABLE IF EXISTS sumar_deseuri CASCADE;
DROP TABLE IF EXISTS sumar_firme CASCADE;
DROP TABLE IF EXISTS vanzari CASCADE;
DROP TABLE IF EXISTS firme CASCADE;

-- Firme (companies/clients)
CREATE TABLE firme (
    id SERIAL PRIMARY KEY,
    name VARCHAR(200) NOT NULL,
    name_normalized VARCHAR(200),
    country VARCHAR(100),
    city VARCHAR(100),
    is_active BOOLEAN DEFAULT true,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(name_normalized)
);
CREATE INDEX idx_firme_name ON firme(name);
CREATE INDEX idx_firme_normalized ON firme(name_normalized);

-- Vanzari (sales - each aviz/delivery note)
CREATE TABLE vanzari (
    id SERIAL PRIMARY KEY,
    firma_id INTEGER REFERENCES firme(id),
    data DATE NOT NULL,
    year INTEGER NOT NULL,
    month INTEGER NOT NULL,
    numar_aviz VARCHAR(100),
    tip_deseu VARCHAR(100),
    cantitate_livrata DECIMAL(14,2),
    pret_achizitie DECIMAL(12,4),
    scazamant_kg DECIMAL(12,2),
    scazamant_ron DECIMAL(12,2),
    cantitate_receptionata DECIMAL(14,2),
    pret_vanzare DECIMAL(12,4),
    valoare_ron DECIMAL(14,2),
    valoare_euro DECIMAL(14,2),
    adaos DECIMAL(14,2),
    transport_ron DECIMAL(12,2),
    adaos_final DECIMAL(14,2),
    serie_factura VARCHAR(50),
    numar_factura VARCHAR(50),
    data_factura DATE,
    observatii TEXT
);
CREATE INDEX idx_vanzari_firma ON vanzari(firma_id);
CREATE INDEX idx_vanzari_date ON vanzari(data);
CREATE INDEX idx_vanzari_year_month ON vanzari(year, month);
CREATE INDEX idx_vanzari_tip_deseu ON vanzari(tip_deseu);

-- Sumar Firme (monthly summary by company)
CREATE TABLE sumar_firme (
    id SERIAL PRIMARY KEY,
    year INTEGER NOT NULL,
    month INTEGER NOT NULL,
    firma_id INTEGER REFERENCES firme(id),
    cantitate_livrata DECIMAL(14,2),
    pret_mediu_achizitie DECIMAL(12,4),
    scazamant_kg DECIMAL(12,2),
    scazamant_ron DECIMAL(12,2),
    cantitate_receptionata DECIMAL(14,2),
    pret_mediu_vanzare DECIMAL(12,4),
    valoare_ron DECIMAL(14,2),
    valoare_euro DECIMAL(14,2),
    transport_ron DECIMAL(12,2),
    adaos DECIMAL(14,2),
    adaos_final DECIMAL(14,2),
    UNIQUE(year, month, firma_id)
);
CREATE INDEX idx_sumar_firme_year_month ON sumar_firme(year, month);
CREATE INDEX idx_sumar_firme_firma ON sumar_firme(firma_id);

-- Sumar Deseuri (monthly summary by waste type)
CREATE TABLE sumar_deseuri (
    id SERIAL PRIMARY KEY,
    year INTEGER NOT NULL,
    month INTEGER NOT NULL,
    tip_deseu VARCHAR(100) NOT NULL,
    cantitate_kg DECIMAL(14,2),
    valoare_ron DECIMAL(14,2),
    adaos_ron DECIMAL(14,2),
    procent_vanzari DECIMAL(10,6),
    procent_profit DECIMAL(10,6),
    UNIQUE(year, month, tip_deseu)
);
CREATE INDEX idx_sumar_deseuri_year_month ON sumar_deseuri(year, month);
CREATE INDEX idx_sumar_deseuri_tip ON sumar_deseuri(tip_deseu);

-- Transporturi Firme (transport costs)
CREATE TABLE transporturi_firme (
    id SERIAL PRIMARY KEY,
    year INTEGER NOT NULL,
    month INTEGER NOT NULL,
    destinatie VARCHAR(200),
    firma_name VARCHAR(200),
    descriere VARCHAR(200),
    suma_fara_tva DECIMAL(12,2),
    tva DECIMAL(10,2),
    total DECIMAL(12,2),
    transportator VARCHAR(100)
);
CREATE INDEX idx_transporturi_year_month ON transporturi_firme(year, month);
""")

conn.commit()
print("    Tabele create!")

# =============================================================================
# COLLECT ALL COMPANY NAMES
# =============================================================================
print("\n[2] Colectare nume firme din toate fisierele...")

all_companies = set()
files_info = [
    ('SITUATIA VANZARILOR 2022_CU TIP DESEU.xlsx', 2022),
    ('SITUATIA VANZARILOR AN 2023.xlsx', 2023),
    ('SITUATIA VANZARILOR AN 2024.xlsx', 2024),
]

for filename, year in files_info:
    filepath = os.path.join(BASE_PATH, filename)
    print(f"    Scanare {filename}...")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    for sheet_name in wb.sheetnames:
        # Skip summary/observatii/transport sheets for company extraction
        if any(x in sheet_name.upper() for x in ['SUMAR', 'OBSERVATII', 'TRANSPORT', 'TOTAL']):
            continue

        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                company = normalize_string(row[0])
                if company and len(company) > 1:
                    all_companies.add(company)

    wb.close()

print(f"    Gasit {len(all_companies)} firme unice")

# Insert companies
print("\n[3] Inserare firme...")
company_id_map = {}  # name_normalized -> id

for company in sorted(all_companies):
    normalized = normalize_company_name(company)
    if normalized and normalized not in company_id_map:
        cur.execute("""
            INSERT INTO firme (name, name_normalized)
            VALUES (%s, %s)
            ON CONFLICT (name_normalized) DO UPDATE SET name = EXCLUDED.name
            RETURNING id
        """, (company, normalized))
        result = cur.fetchone()
        if result:
            company_id_map[normalized] = result[0]

conn.commit()
print(f"    Inserat {len(company_id_map)} firme")

def get_firma_id(company_name):
    """Get firma ID by name"""
    if not company_name:
        return None
    normalized = normalize_company_name(company_name)
    return company_id_map.get(normalized)

# =============================================================================
# IMPORT 2024 DATA (most detailed)
# =============================================================================
print("\n[4] Import date 2024...")

wb = openpyxl.load_workbook(os.path.join(BASE_PATH, 'SITUATIA VANZARILOR AN 2024.xlsx'), data_only=True)

vanzari_2024 = []
sumar_firme_2024 = []
sumar_deseuri_2024 = []

for month_name, month_num in MONTH_MAP.items():
    # Main sheet (vanzari)
    if month_name in wb.sheetnames:
        print(f"    {month_name}...")
        ws = wb[month_name]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            company = normalize_string(row[0])
            firma_id = get_firma_id(company)

            data = parse_date(row[1])
            numar_aviz = normalize_string(row[2])
            cantitate_livrata = safe_decimal(row[4])
            pret_achizitie = safe_decimal(row[5])
            scazamant_kg = safe_decimal(row[6])
            scazamant_ron = safe_decimal(row[7])
            cantitate_receptionata = safe_decimal(row[8])
            pret_vanzare = safe_decimal(row[9])
            valoare_ron = safe_decimal(row[10])
            adaos = safe_decimal(row[11])
            transport_ron = safe_decimal(row[12])
            adaos_final = safe_decimal(row[13])
            serie_factura = normalize_string(row[14]) if len(row) > 14 else None
            numar_factura = normalize_string(row[15]) if len(row) > 15 else None
            data_factura = parse_date(row[16]) if len(row) > 16 else None
            valoare_euro = safe_decimal(row[17]) if len(row) > 17 else None
            observatii = normalize_string(row[18]) if len(row) > 18 else None

            if firma_id and data:
                vanzari_2024.append((
                    firma_id, data, 2024, month_num, numar_aviz, None,
                    cantitate_livrata, pret_achizitie, scazamant_kg, scazamant_ron,
                    cantitate_receptionata, pret_vanzare, valoare_ron, valoare_euro,
                    adaos, transport_ron, adaos_final,
                    serie_factura, numar_factura, data_factura, observatii
                ))

    # Sumar sheet
    sumar_sheet = f'Sumar_{month_name.capitalize()}'
    if sumar_sheet in wb.sheetnames:
        ws = wb[sumar_sheet]

        in_companies_section = False
        in_deseuri_section = False

        for row in ws.iter_rows(values_only=True):
            if not row[0]:
                continue

            row_text = str(row[0]).upper()

            # Detect section
            if 'PE TIP DE CLIENTI' in row_text:
                in_companies_section = True
                in_deseuri_section = False
                continue
            elif 'PE TIP DE DESEURI' in row_text:
                in_companies_section = False
                in_deseuri_section = True
                continue
            elif 'DENUMIRE' in row_text or 'TIP DESEU' in row_text:
                continue
            elif 'GRAND TOTAL' in row_text or 'TOTAL GENERAL' in row_text:
                continue

            if in_companies_section:
                company = normalize_string(row[0])
                firma_id = get_firma_id(company)
                if firma_id:
                    sumar_firme_2024.append((
                        2024, month_num, firma_id,
                        safe_decimal(row[1]),  # cantitate_livrata
                        safe_decimal(row[2]),  # pret_mediu_achizitie
                        safe_decimal(row[3]),  # scazamant_kg
                        safe_decimal(row[4]),  # scazamant_ron
                        safe_decimal(row[5]),  # cantitate_receptionata
                        safe_decimal(row[6]),  # pret_mediu_vanzare
                        safe_decimal(row[7]),  # valoare_ron
                        safe_decimal(row[8]),  # valoare_euro
                        safe_decimal(row[9]),  # transport_ron
                        safe_decimal(row[10]), # adaos
                        safe_decimal(row[11])  # adaos_final
                    ))

            elif in_deseuri_section:
                tip_deseu = normalize_string(row[0])
                if tip_deseu and tip_deseu.upper() not in ['TIP DESEU', 'TOTAL', 'TOTAL GENERAL']:
                    sumar_deseuri_2024.append((
                        2024, month_num, tip_deseu,
                        safe_decimal(row[1]),  # cantitate_kg
                        safe_decimal(row[2]),  # valoare_ron
                        safe_decimal(row[3]),  # adaos_ron
                        safe_decimal(row[4]),  # procent_vanzari
                        safe_decimal(row[5])   # procent_profit
                    ))

wb.close()

# Insert 2024 data
print(f"    Inserare {len(vanzari_2024)} vanzari...")
if vanzari_2024:
    execute_values(cur, """
        INSERT INTO vanzari (firma_id, data, year, month, numar_aviz, tip_deseu,
                            cantitate_livrata, pret_achizitie, scazamant_kg, scazamant_ron,
                            cantitate_receptionata, pret_vanzare, valoare_ron, valoare_euro,
                            adaos, transport_ron, adaos_final,
                            serie_factura, numar_factura, data_factura, observatii)
        VALUES %s
    """, vanzari_2024)

print(f"    Inserare {len(sumar_firme_2024)} sumar firme...")
if sumar_firme_2024:
    execute_values(cur, """
        INSERT INTO sumar_firme (year, month, firma_id, cantitate_livrata, pret_mediu_achizitie,
                                scazamant_kg, scazamant_ron, cantitate_receptionata, pret_mediu_vanzare,
                                valoare_ron, valoare_euro, transport_ron, adaos, adaos_final)
        VALUES %s
        ON CONFLICT (year, month, firma_id) DO UPDATE SET
            cantitate_livrata = EXCLUDED.cantitate_livrata,
            adaos_final = EXCLUDED.adaos_final
    """, sumar_firme_2024)

print(f"    Inserare {len(sumar_deseuri_2024)} sumar deseuri...")
if sumar_deseuri_2024:
    execute_values(cur, """
        INSERT INTO sumar_deseuri (year, month, tip_deseu, cantitate_kg, valoare_ron,
                                  adaos_ron, procent_vanzari, procent_profit)
        VALUES %s
        ON CONFLICT (year, month, tip_deseu) DO UPDATE SET
            cantitate_kg = EXCLUDED.cantitate_kg,
            valoare_ron = EXCLUDED.valoare_ron
    """, sumar_deseuri_2024)

conn.commit()

# =============================================================================
# IMPORT 2023 DATA
# =============================================================================
print("\n[5] Import date 2023...")

wb = openpyxl.load_workbook(os.path.join(BASE_PATH, 'SITUATIA VANZARILOR AN 2023.xlsx'), data_only=True)

vanzari_2023 = []
sumar_firme_2023 = []
sumar_deseuri_2023 = []

for month_name, month_num in MONTH_MAP.items():
    if month_name in wb.sheetnames:
        print(f"    {month_name}...")
        ws = wb[month_name]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            company = normalize_string(row[0])
            firma_id = get_firma_id(company)

            data = parse_date(row[1])
            numar_aviz = normalize_string(row[2])
            cantitate_livrata = safe_decimal(row[4])
            pret_achizitie = safe_decimal(row[5])
            scazamant_kg = safe_decimal(row[6])
            scazamant_ron = safe_decimal(row[7])
            cantitate_receptionata = safe_decimal(row[8])
            pret_vanzare = safe_decimal(row[9])
            valoare_ron = safe_decimal(row[10])
            adaos = safe_decimal(row[11])
            transport_ron = safe_decimal(row[12])
            adaos_final = safe_decimal(row[13])
            observatii = normalize_string(row[14]) if len(row) > 14 else None

            if firma_id and data:
                vanzari_2023.append((
                    firma_id, data, 2023, month_num, numar_aviz, None,
                    cantitate_livrata, pret_achizitie, scazamant_kg, scazamant_ron,
                    cantitate_receptionata, pret_vanzare, valoare_ron, None,
                    adaos, transport_ron, adaos_final,
                    None, None, None, observatii
                ))

    # Sumar sheet
    sumar_sheet = f'Sumar_{month_name.capitalize()}'
    if sumar_sheet in wb.sheetnames:
        ws = wb[sumar_sheet]

        in_companies_section = False
        in_deseuri_section = False

        for row in ws.iter_rows(values_only=True):
            if not row[0]:
                continue

            row_text = str(row[0]).upper()

            if 'PE TIP DE CLIENTI' in row_text:
                in_companies_section = True
                in_deseuri_section = False
                continue
            elif 'PE TIP DE DESEURI' in row_text:
                in_companies_section = False
                in_deseuri_section = True
                continue
            elif 'DENUMIRE' in row_text or 'TIP DESEU' in row_text:
                continue
            elif 'GRAND TOTAL' in row_text or 'TOTAL GENERAL' in row_text:
                continue

            if in_companies_section:
                company = normalize_string(row[0])
                firma_id = get_firma_id(company)
                if firma_id:
                    sumar_firme_2023.append((
                        2023, month_num, firma_id,
                        safe_decimal(row[1]), safe_decimal(row[2]), safe_decimal(row[3]),
                        safe_decimal(row[4]), safe_decimal(row[5]), safe_decimal(row[6]),
                        safe_decimal(row[7]), safe_decimal(row[8]), safe_decimal(row[9]),
                        safe_decimal(row[10]), safe_decimal(row[11])
                    ))

            elif in_deseuri_section:
                tip_deseu = normalize_string(row[0])
                if tip_deseu and tip_deseu.upper() not in ['TIP DESEU', 'TOTAL', 'TOTAL GENERAL']:
                    sumar_deseuri_2023.append((
                        2023, month_num, tip_deseu,
                        safe_decimal(row[1]), safe_decimal(row[2]), safe_decimal(row[3]),
                        safe_decimal(row[4]), safe_decimal(row[5])
                    ))

wb.close()

print(f"    Inserare {len(vanzari_2023)} vanzari...")
if vanzari_2023:
    execute_values(cur, """
        INSERT INTO vanzari (firma_id, data, year, month, numar_aviz, tip_deseu,
                            cantitate_livrata, pret_achizitie, scazamant_kg, scazamant_ron,
                            cantitate_receptionata, pret_vanzare, valoare_ron, valoare_euro,
                            adaos, transport_ron, adaos_final,
                            serie_factura, numar_factura, data_factura, observatii)
        VALUES %s
    """, vanzari_2023)

if sumar_firme_2023:
    execute_values(cur, """
        INSERT INTO sumar_firme (year, month, firma_id, cantitate_livrata, pret_mediu_achizitie,
                                scazamant_kg, scazamant_ron, cantitate_receptionata, pret_mediu_vanzare,
                                valoare_ron, valoare_euro, transport_ron, adaos, adaos_final)
        VALUES %s
        ON CONFLICT (year, month, firma_id) DO NOTHING
    """, sumar_firme_2023)

if sumar_deseuri_2023:
    execute_values(cur, """
        INSERT INTO sumar_deseuri (year, month, tip_deseu, cantitate_kg, valoare_ron,
                                  adaos_ron, procent_vanzari, procent_profit)
        VALUES %s
        ON CONFLICT (year, month, tip_deseu) DO NOTHING
    """, sumar_deseuri_2023)

conn.commit()

# =============================================================================
# IMPORT 2022 DATA (has tip_deseu!)
# =============================================================================
print("\n[6] Import date 2022 (cu tip deseu)...")

wb = openpyxl.load_workbook(os.path.join(BASE_PATH, 'SITUATIA VANZARILOR 2022_CU TIP DESEU.xlsx'), data_only=True)

vanzari_2022 = []

for month_name, month_num in MONTH_MAP.items():
    if month_name in wb.sheetnames:
        print(f"    {month_name}...")
        ws = wb[month_name]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            company = normalize_string(row[0])
            firma_id = get_firma_id(company)

            data = parse_date(row[1])
            numar_aviz = normalize_string(row[2])
            cantitate_livrata = safe_decimal(row[4])
            pret_achizitie = safe_decimal(row[5])
            scazamant_kg = safe_decimal(row[6])
            scazamant_ron = safe_decimal(row[7])
            cantitate_receptionata = safe_decimal(row[8])
            pret_vanzare = safe_decimal(row[9])
            valoare_ron = safe_decimal(row[10])
            adaos = safe_decimal(row[11])
            transport_ron = safe_decimal(row[12]) if len(row) > 12 else None
            adaos_final = safe_decimal(row[13]) if len(row) > 13 else None
            tip_deseu = normalize_string(row[14]) if len(row) > 14 else None

            if firma_id and data:
                vanzari_2022.append((
                    firma_id, data, 2022, month_num, numar_aviz, tip_deseu,
                    cantitate_livrata, pret_achizitie, scazamant_kg, scazamant_ron,
                    cantitate_receptionata, pret_vanzare, valoare_ron, None,
                    adaos, transport_ron, adaos_final,
                    None, None, None, None
                ))

wb.close()

print(f"    Inserare {len(vanzari_2022)} vanzari...")
if vanzari_2022:
    execute_values(cur, """
        INSERT INTO vanzari (firma_id, data, year, month, numar_aviz, tip_deseu,
                            cantitate_livrata, pret_achizitie, scazamant_kg, scazamant_ron,
                            cantitate_receptionata, pret_vanzare, valoare_ron, valoare_euro,
                            adaos, transport_ron, adaos_final,
                            serie_factura, numar_factura, data_factura, observatii)
        VALUES %s
    """, vanzari_2022)

conn.commit()

# =============================================================================
# IMPORT TRANSPORTURI
# =============================================================================
print("\n[7] Import transporturi...")

for filename, year in [('SITUATIA VANZARILOR AN 2024.xlsx', 2024), ('SITUATIA VANZARILOR AN 2023.xlsx', 2023)]:
    wb = openpyxl.load_workbook(os.path.join(BASE_PATH, filename), data_only=True)

    if 'transporturi' in wb.sheetnames:
        print(f"    {year} transporturi...")
        ws = wb['transporturi']

        transporturi = []
        current_month = 1

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not any(row):
                continue

            # Check if this is a month header
            if row[0] and str(row[0]).upper() in MONTH_MAP:
                current_month = MONTH_MAP[str(row[0]).upper()]
                continue

            destinatie = normalize_string(row[0])
            firma_name = normalize_string(row[1])
            descriere = normalize_string(row[2])
            suma_fara_tva = safe_decimal(row[3])
            tva = safe_decimal(row[4])
            total = safe_decimal(row[5])
            transportator = normalize_string(row[6]) if len(row) > 6 else None

            if destinatie or firma_name:
                transporturi.append((
                    year, current_month, destinatie, firma_name, descriere,
                    suma_fara_tva, tva, total, transportator
                ))

        if transporturi:
            execute_values(cur, """
                INSERT INTO transporturi_firme (year, month, destinatie, firma_name, descriere,
                                               suma_fara_tva, tva, total, transportator)
                VALUES %s
            """, transporturi)
            print(f"      Inserat {len(transporturi)} transporturi")

    wb.close()

conn.commit()

# =============================================================================
# FINAL STATISTICS
# =============================================================================
print("\n" + "=" * 70)
print("STATISTICI FINALE - FIRME")
print("=" * 70)

cur.execute("SELECT COUNT(*) FROM firme")
print(f"  Firme: {cur.fetchone()[0]}")

cur.execute("SELECT COUNT(*) FROM vanzari")
print(f"  Vanzari (avize): {cur.fetchone()[0]}")

cur.execute("SELECT COUNT(*) FROM sumar_firme")
print(f"  Sumar lunar firme: {cur.fetchone()[0]}")

cur.execute("SELECT COUNT(*) FROM sumar_deseuri")
print(f"  Sumar lunar deseuri: {cur.fetchone()[0]}")

cur.execute("SELECT COUNT(*) FROM transporturi_firme")
print(f"  Transporturi: {cur.fetchone()[0]}")

cur.execute("SELECT MIN(data), MAX(data) FROM vanzari")
date_range = cur.fetchone()
print(f"  Perioada: {date_range[0]} - {date_range[1]}")

cur.execute("SELECT SUM(valoare_ron) FROM vanzari")
total = cur.fetchone()[0]
print(f"  Valoare totala vanzari: {total:,.2f} RON")

cur.execute("SELECT SUM(adaos_final) FROM vanzari WHERE adaos_final IS NOT NULL")
profit = cur.fetchone()[0]
if profit:
    print(f"  Profit total (adaos final): {profit:,.2f} RON")

print("\n--- Top 5 firme (dupa valoare) ---")
cur.execute("""
    SELECT f.name, SUM(v.valoare_ron) as total_ron, COUNT(*) as nr_avize
    FROM vanzari v
    JOIN firme f ON v.firma_id = f.id
    GROUP BY f.id, f.name
    ORDER BY total_ron DESC
    LIMIT 5
""")
for row in cur.fetchall():
    print(f"  {row[0]}: {row[1]:,.2f} RON ({row[2]} avize)")

print("\n--- Top 5 tipuri deseuri (dupa valoare, 2024) ---")
cur.execute("""
    SELECT tip_deseu, SUM(valoare_ron) as total_ron, SUM(cantitate_kg) as total_kg
    FROM sumar_deseuri
    WHERE year = 2024
    GROUP BY tip_deseu
    ORDER BY total_ron DESC
    LIMIT 5
""")
for row in cur.fetchall():
    print(f"  {row[0]}: {row[1]:,.2f} RON ({row[2]:,.0f} kg)")

cur.close()
conn.close()

print("\n" + "=" * 70)
print("IMPORT COMPLET!")
print("=" * 70)
